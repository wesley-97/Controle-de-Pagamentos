import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class FinancialControlApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Controle Financeiro de Obra')
        self.geometry('1100x1000')
        self.configure(bg='#f6f6f6')

        self.start_date = datetime.now()
        self.payments = []
        self.editing_index = None

        self.create_widgets()

    def create_widgets(self):
        frame_inputs = ttk.LabelFrame(self, text="Lançamento de Pagamento")
        frame_inputs.pack(fill='x', padx=20, pady=10)

        ttk.Label(frame_inputs, text='Nome do Pagador:').grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.entry_payer = ttk.Combobox(frame_inputs, values=['Wesley', 'Beatriz'])
        self.entry_payer.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        self.entry_payer.current(0)

        ttk.Label(frame_inputs, text='Finalidade do Pagamento:').grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.entry_purpose = ttk.Combobox(frame_inputs, values=['Entrada', 'Evolução de Obra', 'Taxas', 'Financiamento'])
        self.entry_purpose.grid(row=0, column=3, sticky='w', padx=5, pady=5)
        self.entry_purpose.current(0)

        ttk.Label(frame_inputs, text='Valor Pago (R$):').grid(row=0, column=4, sticky='w', padx=5, pady=5)
        self.entry_value = ttk.Entry(frame_inputs, width=12)
        self.entry_value.grid(row=0, column=5, sticky='w', padx=5, pady=5)

        ttk.Label(frame_inputs, text='Data (dd/mm/aaaa):').grid(row=0, column=6, sticky='w', padx=5, pady=5)
        self.entry_date = ttk.Entry(frame_inputs, width=12)
        self.entry_date.grid(row=0, column=7, sticky='w', padx=5, pady=5)
        self.entry_date.insert(0, date.today().strftime('%d/%m/%Y'))

        ttk.Label(frame_inputs, text='Obs:').grid(row=0, column=8, sticky='w', padx=5, pady=5)
        self.entry_note = ttk.Entry(frame_inputs, width=30)
        self.entry_note.grid(row=0, column=9, sticky='w', padx=5, pady=5)

        self.button_add = ttk.Button(frame_inputs, text='Adicionar Pagamento', command=self.add_payment)
        self.button_add.grid(row=0, column=10, padx=10, pady=5)
        self.button_save = ttk.Button(frame_inputs, text='Salvar Edição', command=self.save_edit)

        frame_table = ttk.LabelFrame(self, text="Pagamentos Registrados")
        frame_table.pack(fill='both', expand=True, padx=20, pady=5)

        self.tree = ttk.Treeview(frame_table, columns=('Pagador', 'Finalidade', 'Valor', 'Data', 'Observação'), show='headings', height=8)
        self.tree.heading('Pagador', text='Pagador')
        self.tree.heading('Finalidade', text='Finalidade')
        self.tree.heading('Valor', text='Valor (R$)')
        self.tree.heading('Data', text='Data')
        self.tree.heading('Observação', text='Observação')

        self.tree.column('Pagador', width=120, anchor='center')
        self.tree.column('Finalidade', width=150, anchor='center')
        self.tree.column('Valor', width=100, anchor='e')
        self.tree.column('Data', width=95, anchor='center')
        self.tree.column('Observação', width=250, anchor='w')

        self.tree.pack(fill='both', expand=True, padx=5, pady=5)

        btn_frame = ttk.Frame(frame_table)
        btn_frame.pack(pady=8)
        ttk.Button(btn_frame, text="Apagar Selecionado", command=self.delete_selected).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Editar Selecionado", command=self.edit_selected).pack(side='left', padx=5)

        frame_dash = ttk.LabelFrame(self, text="Dashboard de Controle")
        frame_dash.pack(fill='x', padx=20, pady=15)

        ttk.Label(frame_dash, text='Data prevista para conclusão (dd/mm/aaaa):').pack(anchor='center', pady=2)
        self.entry_finish_date = ttk.Entry(frame_dash, width=14)
        self.entry_finish_date.pack(anchor='center', pady=2)
        default_finish_date = date.today().replace(year=date.today().year+2)
        self.entry_finish_date.insert(0, default_finish_date.strftime('%d/%m/%Y'))

        self.label_time_left = ttk.Label(frame_dash, text='Tempo restante para conclusão: -- meses', font=('Helvetica', 12, 'bold'))
        self.label_time_left.pack(anchor='center', pady=5)

        dash_btns = ttk.Frame(frame_dash)
        dash_btns.pack(anchor='center')
        ttk.Button(dash_btns, text='Atualizar Dashboard', command=self.update_dashboard).pack(side='left', padx=5)
        ttk.Button(dash_btns, text='Exportar Relatório Excel', command=self.export_excel).pack(side='left', padx=5)

        frame_graph = tk.Frame(self, bg='#f6f6f6')
        frame_graph.pack(fill='x', anchor='center', padx=20, pady=10)

        self.fig, self.axs = plt.subplots(2, 1, figsize=(6, 14))
        self.fig.subplots_adjust(hspace=1.5)
        self.canvas = FigureCanvasTkAgg(self.fig, master=frame_graph)
        self.canvas.get_tk_widget().pack()

        self.update_dashboard()

    def add_payment(self):
        if self.editing_index is not None:
            messagebox.showwarning('Atenção', 'Salve ou cancele a edição antes de adicionar novo pagamento.')
            return
        payer = self.entry_payer.get()
        purpose = self.entry_purpose.get()
        date_str = self.entry_date.get().strip()
        note_str = self.entry_note.get().strip()
        try:
            value = float(self.entry_value.get())
        except ValueError:
            messagebox.showerror('Erro', 'Por favor insira um valor numérico válido.')
            return

        self.payments.append({'payer': payer, 'purpose': purpose, 'value': value, 'date': date_str, 'note': note_str})
        self.clear_form()
        self.update_dashboard()

    def edit_selected(self):
        selected = self.tree.selection()
        if not selected or len(selected) != 1:
            messagebox.showwarning("Aviso", "Selecione UM lançamento para editar.")
            return
        item = selected[0]
        values = self.tree.item(item, "values")
        for idx, p in enumerate(self.payments):
            if (p['payer'], p['purpose'], f"{p['value']:.2f}", p['date'], p['note']) == values:
                self.editing_index = idx
                self.entry_payer.set(p['payer'])
                self.entry_purpose.set(p['purpose'])
                self.entry_value.delete(0, tk.END)
                self.entry_value.insert(0, str(p['value']))
                self.entry_date.delete(0, tk.END)
                self.entry_date.insert(0, p['date'])
                self.entry_note.delete(0, tk.END)
                self.entry_note.insert(0, p['note'])
                self.button_add.grid_remove()
                self.button_save.grid(row=0, column=10, padx=10, pady=5)
                break

    def save_edit(self):
        if self.editing_index is None:
            return
        payer = self.entry_payer.get()
        purpose = self.entry_purpose.get()
        date_str = self.entry_date.get().strip()
        note_str = self.entry_note.get().strip()
        try:
            value = float(self.entry_value.get())
        except ValueError:
            messagebox.showerror('Erro', 'Por favor insira um valor numérico válido.')
            return
        self.payments[self.editing_index] = {'payer': payer, 'purpose': purpose, 'value': value, 'date': date_str, 'note': note_str}
        self.editing_index = None
        self.button_save.grid_remove()
        self.button_add.grid(row=0, column=10, padx=10, pady=5)
        self.clear_form()
        self.update_dashboard()

    def clear_form(self):
        self.entry_value.delete(0, tk.END)
        self.entry_date.delete(0, tk.END)
        self.entry_note.delete(0, tk.END)
        self.entry_payer.current(0)
        self.entry_purpose.current(0)
        self.entry_date.insert(0, date.today().strftime('%d/%m/%Y'))

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um lançamento para apagar.")
            return
        for item in selected:
            values = self.tree.item(item, "values")
            for i, p in enumerate(self.payments):
                if (p['payer'], p['purpose'], f"{p['value']:.2f}", p['date'], p['note']) == values:
                    del self.payments[i]
                    break
            self.tree.delete(item)
        self.editing_index = None
        self.button_save.grid_remove()
        self.button_add.grid(row=0, column=10, padx=10, pady=5)
        self.clear_form()
        self.update_dashboard()

    def update_dashboard(self):
        try:
            finish_date = datetime.strptime(self.entry_finish_date.get().strip(), '%d/%m/%Y')
            today = datetime.now()
            delta = finish_date - today
            months_left = max(0, delta.days // 30)
        except Exception:
            months_left = '--'
        self.label_time_left.config(text=f'Tempo restante para conclusão: {months_left} meses')

        totals_by_payer = {'Wesley': 0, 'Beatriz': 0}
        totals_by_purpose = {'Entrada': 0, 'Evolução de Obra': 0, 'Taxas': 0, 'Financiamento': 0}

        for p in self.payments:
            totals_by_payer[p['payer']] += p['value']
            totals_by_purpose[p['purpose']] += p['value']

        self.axs[0].clear()
        self.axs[1].clear()

        for i, v in enumerate(totals_by_payer.values()):
            valor_formatado = f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            self.axs[0].bar(totals_by_payer.keys(), totals_by_payer.values(), color=['steelblue', 'orange'])
            self.axs[0].set_title('Total por pagador', fontsize=12, weight='bold', pad=30)
            self.axs[0].axes.get_yaxis().set_visible(False)
            self.axs[0].spines['top'].set_visible(False)
            self.axs[0].spines['right'].set_visible(False)
            self.axs[0].spines['left'].set_visible(False)
            self.axs[0].spines['bottom'].set_visible(False)
            self.axs[0].text(i, v, valor_formatado, ha='center', va='bottom', fontsize=13, fontweight='bold')

        for i, v in enumerate(totals_by_purpose.values()):
            valor_formatado = f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            self.axs[1].bar(totals_by_purpose.keys(), totals_by_purpose.values(), color='forestgreen')
            self.axs[1].set_title('Total por finalidade', fontsize=12, weight='bold', pad=30)
            self.axs[1].axes.get_yaxis().set_visible(False)
            self.axs[1].spines['top'].set_visible(False)
            self.axs[1].spines['right'].set_visible(False)
            self.axs[1].spines['left'].set_visible(False)
            self.axs[1].spines['bottom'].set_visible(False)
            self.axs[1].text(i, v, valor_formatado, ha='center', va='bottom', fontsize=13, fontweight='bold')

        self.fig.subplots_adjust(hspace=1.5)
        self.canvas.draw()

        # Ordenar pagamentos por data (crescente)
        self.payments.sort(key=lambda x: datetime.strptime(x['date'], "%d/%m/%Y"))
        # Atualizar TreeView
        self.tree.delete(*self.tree.get_children())
        for p in self.payments:
            self.tree.insert('', 'end', values=(p['payer'], p['purpose'], f"{p['value']:.2f}", p['date'], p['note']))

    def export_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx',
            filetypes=[('Excel Files', '*.xlsx')],
            title='Salvar relatório como...')
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório Financeiro"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            headers = ['Pagador', 'Finalidade', 'Valor (R$)', 'Data', 'Observação']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for row_idx, p in enumerate(self.payments, 2):
                ws.cell(row=row_idx, column=1).value = p['payer']
                ws.cell(row=row_idx, column=2).value = p['purpose']
                ws.cell(row=row_idx, column=3).value = p['value']
                ws.cell(row=row_idx, column=4).value = p['date']
                ws.cell(row=row_idx, column=5).value = p['note']
                ws.cell(row=row_idx, column=3).number_format = 'R$ #,##0.00'

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 35

            wb.save(file_path)
            messagebox.showinfo('Sucesso', f'Relatório exportado como {file_path}')
        except Exception as e:
            messagebox.showerror('Erro', f'Não foi possível salvar o relatório: {e}')

if __name__ == '__main__':
    app = FinancialControlApp()
    app.mainloop()